# Sanitized Demo Script
# This script is provided as a demonstration using synthetic data only.
# Do not use it with proprietary/internal datasets without authorization.

#!/usr/bin/env python3
import csv
import socket
import subprocess
import platform
import ipaddress
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PING_COUNT = 2
PING_TIMEOUT_SEC = 2
STOP_THIRD_OCTET = 255  # demo default (no early stop)


# ---------- helpers ----------
def get_third_octet(ip_str: str):
    ip = ipaddress.ip_address(ip_str)
    if ip.version != 4:
        return None
    return int(str(ip).split(".")[2])


def dns_resolve(name: str):
    try:
        infos = socket.getaddrinfo(name, None)
        ips = sorted({info[4][0] for info in infos})
        return True, ";".join(ips), ""
    except Exception as e:
        return False, "", str(e)


def ping(host: str):
    system = platform.system().lower()
    if system == "windows":
        cmd = ["ping", "-n", str(PING_COUNT), "-w", str(PING_TIMEOUT_SEC * 1000), host]
        hard_timeout = PING_TIMEOUT_SEC + 2
    else:
        cmd = ["ping", "-c", str(PING_COUNT), "-W", str(PING_TIMEOUT_SEC), host]
        hard_timeout = PING_TIMEOUT_SEC + 2

    try:
        r = subprocess.run(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
            text=True,
            timeout=hard_timeout
        )
        return (r.returncode == 0), (r.stderr.strip() if r.returncode != 0 else "")
    except Exception as e:
        return False, str(e)


def disposition_for(result, dns_resolved, ping_succeeded):
    """
    Matches your final mapping:
      OK -> Keep / Active
      DNS_FAIL -> Needs DNS Review
      NO_RESPONSE + DNS TRUE + ping FALSE -> Needs Manual Verify
      NO_RESPONSE + DNS not TRUE + ping FALSE -> Candidate Stale
    """
    # normalize booleans / strings
    dns_true = (str(dns_resolved).strip().lower() == "true")
    ping_true = (str(ping_succeeded).strip().lower() == "true")

    if result == "OK":
        return "Keep / Active"
    if result == "DNS_FAIL":
        return "Needs DNS Review"
    if result == "NO_RESPONSE" and dns_true and not ping_true:
        return "Needs Manual Verify"
    if result == "NO_RESPONSE" and (not dns_true) and not ping_true:
        return "Candidate Stale"
    # safe fallback
    return "Needs Manual Verify"


def result_fill(result, dns_resolved, ping_succeeded):
    """
    Color logic aligned to your sheet intent:
      OK -> green
      DNS_FAIL -> orange
      NO_RESPONSE -> red, BUT if DNS TRUE + ping FALSE -> yellow (partial signal)
    """
    dns_true = (str(dns_resolved).strip().lower() == "true")
    ping_true = (str(ping_succeeded).strip().lower() == "true")

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    orange = PatternFill("solid", fgColor="F8CBAD")
    red = PatternFill("solid", fgColor="FFC7CE")

    if result == "OK":
        return green
    if result == "DNS_FAIL":
        return orange
    if result == "NO_RESPONSE":
        if dns_true and not ping_true:
            return yellow
        return red
    return None


def write_legend(ws, start_row=2, col_prefix=9, col_expl=10):
    """
    Writes the legend block like your current results_colorcoded.xlsx:
    Prefix in col I (9), Explanation in col J (10).
    """
    legend_rows = [
        ("Legend", ""),
        ("●OK rows", "OK – validated"),
        ("● Partial rows (DNS ok, ping fails):", "PARTIAL – DNS resolves, ICMP blocked"),
        ("● Partial rows (Ping ok, DNS fails):", "PARTIAL – Host reachable, DNS lookup failed"),
        ("● DNS_FAIL", "DNS_FAIL – DNS lookup error (Name not found)"),
        ("● NO_RESPONSE", "NO_SIGNAL – No ping response, DNS unavailable"),
    ]

    r = start_row
    for p, e in legend_rows:
        ws.cell(r, col_prefix).value = p
        ws.cell(r, col_expl).value = e
        r += 1

    # Style headers for legend columns
    ws.cell(1, col_prefix).value = "Prefix"
    ws.cell(1, col_expl).value = "Explanation"
    ws.cell(1, col_prefix).font = Font(bold=True)
    ws.cell(1, col_expl).font = Font(bold=True)


def main(input_csv, output_xlsx):
    checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(input_csv, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    wb = Workbook()
    ws = wb.active
    ws.title = "results_colorcoded"

    # Match your sheet layout + add Disposition
    headers = [
        "TargetUsed", "DnsResolved", "ResolvedIPs", "PingSucceeded",
        "Reason", "Result", "Disposition", "CheckedAt"
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = h
        ws.cell(1, i).font = Font(bold=True)

    # spacer column H -> leave blank to match your layout feel
    # legend starts at I/J
    write_legend(ws, start_row=2, col_prefix=9, col_expl=10)

    out_row = 2
    for row in rows:
        ip = (row.get("IP Address") or row.get("TargetUsed") or "").strip()
        name_field = (row.get("Name") or "").strip()

        # hard stop based on 3rd octet
        if ip:
            try:
                third = get_third_octet(ip)
                if third is not None and third >= STOP_THIRD_OCTET:
                    break
            except ValueError:
                pass

        domains = [d.strip() for d in name_field.split(",") if d.strip()] if name_field else []

        # ping target: prefer IP if present, else first domain
        ping_target = ip if ip else (domains[0] if domains else "")

        # no inputs
        if not ip and not domains:
            dns_ok = ""
            resolved_ips = ""
            ping_ok = ""
            reason = "No IP Address or Name provided"
            result = "SKIP"
        else:
            dns_err = ""
            resolved_pairs = []
            dns_ok = "N/A"

            if domains:
                dns_ok = False
                for domain in domains:
                    ok, ips, err = dns_resolve(domain)
                    if ok:
                        dns_ok = True
                        resolved_pairs.append(f"{domain}:{ips}")
                        if not ip:
                            ping_target = domain
                        break
                    else:
                        dns_err = err

            resolved_ips = "; ".join(resolved_pairs)

            ping_ok, ping_err = ping(ping_target)

            reasons = []
            if dns_err:
                reasons.append(f"DNS:{dns_err}")
            if ping_err and not ping_ok:
                reasons.append(f"PING:{ping_err}")
            reason = "; ".join(reasons)

            if domains and dns_ok is False:
                result = "DNS_FAIL"
            elif ping_ok:
                result = "OK"
            else:
                result = "NO_RESPONSE"

        disp = disposition_for(result, dns_ok, ping_ok if isinstance(ping_ok, bool) else str(ping_ok))

        # write row
        values = [
            ping_target,
            dns_ok,
            resolved_ips,
            ping_ok,
            reason,
            result,
            disp,
            checked_at
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(out_row, c).value = v

        # color Result cell
        fill = result_fill(result, dns_ok, ping_ok)
        if fill:
            ws.cell(out_row, 6).fill = fill  # Result column is 6
            ws.cell(out_row, 6).font = Font(bold=True)

        out_row += 1

    # basic formatting
    ws.freeze_panes = "A2"
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.column_dimensions["C"].width = 60  # ResolvedIPs can be long
    ws.column_dimensions["E"].width = 45  # Reason can be long

    wb.save(output_xlsx)
    print(f"Wrote Excel output to {output_xlsx}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Check IP/domain DNS + Ping and output results_colorcoded-style XLSX"
    )
    parser.add_argument("input_csv", help="Input CSV with 'IP Address' and/or 'Name'")
    parser.add_argument(
        "-o", "--output",
        default=f"results_colorcoded_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Output XLSX file"
    )
    args = parser.parse_args()
    main(args.input_csv, args.output)
